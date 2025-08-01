from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import LandPlot
from .forms import LeaseAgreementForm, LeaseFilterForm, ExcelImportForm
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse, HttpResponseRedirect
from datetime import datetime
from django.db.models import Q, F
from .filters import LandPlotFilter
from simple_history.utils import update_change_reason
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views import View
from django.template.loader import render_to_string
from django.utils.html import format_html
import json
from django.urls import reverse
import requests
import pprint
from urllib.parse import urlencode
from django.utils.http import url_has_allowed_host_and_scheme
from django.core.serializers import serialize
from django.contrib.gis.geos import Polygon
from django.contrib.gis.db.models.functions import Transform
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .serializers import LandPlotSerializer
from django.contrib.gis.serializers.geojson import Serializer
import pandas as pd
import os
from django.conf import settings
from openpyxl import Workbook


# Create your views here.
# Головна сторінка
def index(request):
    return render(request, 'land/index.html', {'title': 'Головна сторінка'})

# сторінка з таблицею про оренду
def zem_lease(request):
    land_plots = LandPlot.objects.filter(status='2')
    search_query = request.GET.get('search', '')

    if search_query:
        land_plots = land_plots.filter(
            Q(cadastr_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(area__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(interest__icontains=search_query) |
            Q(assessment__icontains=search_query) |
            Q(rent_start__icontains=search_query) |
            Q(rent_end__icontains=search_query)
        )

    land_filter = LeaseFilterForm(request.GET, queryset=land_plots)
    filtered_qs = land_filter.qs.order_by('id')  # Додаємо сортування
    paginator = Paginator(filtered_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('land/zem_lease_results.html', {
            'leases': page_obj,
            'page_obj': page_obj,
        })
        return JsonResponse({'html': html})

    return render(request, 'land/zem_lease.html', {
        'title': 'Договора оренди',
        'filter': land_filter,
        'leases': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
    })

 # додавання в таблицю інформацію про оренду
def zem_leave_add(request):
    if request.method == 'POST':
        form = LeaseAgreementForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Договір успішно доданий!')
            return redirect('land')  # назад до списку
    else:
        form = LeaseAgreementForm()
    return render(request, 'land/zem_lease_add.html', {'form': form, 'title': 'Новий договір'})

# редагування договору
def zem_lease_edit(request, pk):
    lease = get_object_or_404(LandPlot, pk=pk)
    query_params = request.GET.dict()
    next_url = request.GET.get('next')

    # Захист: перевірка, щоб next не було зовнішнім посиланням
    if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        next_url = reverse('zem_lease_result')
        if query_params:
            next_url += '?' + urlencode(query_params)

    if request.method == "POST":
        form = LeaseAgreementForm(request.POST, instance=lease)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Внесено зміни успішно!')
            return redirect(next_url)
    else:
        form = LeaseAgreementForm(instance=lease)

    return render(request, 'land/zem_lease_edit.html', {
        'form': form,
        'title': 'Редагувати запис',
        'lease': lease,
        'query_params': urlencode(query_params),
    })

#  фільтр
def zem_lease_filter(request):
    land_plots = LandPlot.objects.all()
    land_filter = LeaseFilterForm(request.GET, queryset=land_plots)
    paginator = Paginator(land_filter.qs.order_by('id'), 50)  # Додаємо сортування
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    print(f"GET-параметри в land_export: {request.GET}")

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('land/zem_lease_results.html', {
            'leases': page_obj,
            'page_obj': page_obj,
        })
        return JsonResponse({'html': html})

    return render(request, 'land/zem_lease.html', {
        'filter': land_filter,
        'leases': page_obj,
        'page_obj': page_obj,
    })
# видалення договору
def zem_lease_delete(request, pk):
    lease = get_object_or_404(LandPlot, pk=pk)
    lease.delete()
    messages.success(request, '❌ Земельна ділянка успішно видалена!')

    # Повертаємось на сторінку, з якої прийшов запит (з фільтрами і сторінкою)
    next_url = request.GET.get('next')
    if next_url:
        return HttpResponseRedirect(next_url)
    return redirect('zem_lease_result')

# історія змін
# порівняння рядків
def get_changes(prev, curr):
    changes = []
    for field in curr._meta.fields:
        field_name = field.name
        # Пропускаємо службові поля
        if field_name in ['id', 'history_id', 'history_date', 'history_user', 'history_type']:
            continue
        old = getattr(prev, field_name, None)
        new = getattr(curr, field_name, None)
        if old != new:
            verbose_name = field.verbose_name.capitalize()
            changes.append((verbose_name, old, new))
    return changes

def zem_lease_history(request, pk):
    land_plot = get_object_or_404(LandPlot, pk=pk)
    history = land_plot.history.all().order_by('-history_date')

    history_with_changes = []
    for i in range(len(history) - 1):
        current = history[i]
        previous = history[i + 1]
        changes = get_changes(previous, current)
        history_with_changes.append((current, changes))

    # Додаємо перший запис (створення)
    if history:
        history_with_changes.append((history.last(), []))

    return render(request, 'land/zem_lease_history.html', {
        'history': history_with_changes
    })

# імпортування з Excel

# Перетворення значення категорії в текст
def get_category_key(display_name):
    for key, name in LandPlot.CATEGORY_CHOICES:
        if name == display_name:
            return key
    return None

# Перетворення значення цільового призначення в текст
def get_destination_key(display_name):
    for key, name in LandPlot.DESTINATION_CHOICES:
        if name == display_name:
            return key
    return None

# Перетворення значення угіддя в текст
def get_land_key(display_name):
    for key, name in LandPlot.LAND_CHOICES:
        if name == display_name:
            return key
    return None
# Перетворення значення реєстрації договору в текст
def get_registered_key(display_name):
    for key, name in LandPlot.REGISTER_CHOICES:
        if name == display_name:
            return key
    return None

def get_status_key(display_name):
    for key, name in LandPlot.STATUS_CHOICES:
        if name == display_name:
            return key
    return None

def get_gidr_key(display_name):
    for key, name in LandPlot.GIDR_CHOICES:
        if name == display_name:
            return key
    return None

# превью перед імпортом
def preview_excel_import(request):
    if request.method == "POST" and "preview" in request.POST:
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["file"]
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                preview_data = []

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    def parse_date(value):
                        if isinstance(value, datetime):
                            return value.date().isoformat()
                        elif isinstance(value, str):
                            try:
                                return datetime.strptime(value.strip(), "%d.%m.%Y").date().isoformat()
                            except ValueError:
                                return None
                        return None

                    preview_data.append({
                        "cadastr_number": row[1],
                        "location": row[2],
                        "area": row[3],
                        "owner_name": row[4],
                        "category": row[5],
                        "destination": row[6],
                        "rent_start": parse_date(row[7]),
                        "rent_end": parse_date(row[8]),
                        "interest": row[9],
                        "assessment": row[10],
                        "land": row[11],
                        "registered": row[12],
                        "notes": row[13],
                    })

                request.session["preview_data"] = json.dumps(preview_data, default=str)

                return render(request, "land/preview_import.html", {
                    "preview_data": preview_data
                })

            except Exception as e:
                messages.error(request, f"Помилка при читанні Excel: {str(e)}")

    elif request.method == "POST" and "confirm" in request.POST:
        preview_data = json.loads(request.session.get("preview_data", "[]"))

        for item in preview_data:
            LandPlot.objects.create(
                cadastr_number=item["cadastr_number"],
                location=item["location"],
                area=item["area"],
                owner_name=item["owner_name"],
                category=get_category_key(item["category"]),
                destination=get_destination_key(item["destination"]),
                rent_start=item["rent_start"],
                rent_end=item["rent_end"],
                interest=item["interest"],
                assessment=item["assessment"],
                land=get_land_key(item["land"]),
                registered=get_registered_key(item["registered"]),
                notes=item["notes"]
            )
        messages.success(request, f"Імпортовано {len(preview_data)} записів.")
        return redirect("import_landplots_excel")

    else:
        form = ExcelImportForm()

    return render(request, "land/import_excel.html", {"form": form})


# сам імпорт з excel
def import_excel(request):
    if request.method == "POST":
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["file"]
            errors = []  # зберігатимемо помилки тут

            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                def parse_date(value):
                    if isinstance(value, datetime):
                        return value.date()
                    elif isinstance(value, str):
                        try:
                            return datetime.strptime(value.strip(), "%d.%m.%Y").date()
                        except ValueError:
                            return None
                    return None

                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        cadastr_number = row[1]
                        location = row[2]
                        area = row[3]
                        coordinates = row[4]
                        status = row[5]
                        owner_name = row[6]
                        category = row[7]
                        destination = row[8]
                        rent_start = parse_date(row[9])
                        rent_end = parse_date(row[10])
                        interest = row[11]
                        assessment = row[12]
                        land = row[13]
                        registered = row[14]
                        notes = row[15]
                        pass_number = row[16]
                        pass_name = row[17]
                        pass_area = row[18]
                        pass_volume = row[19]
                        pass_depth = row[20]
                        date_approval = parse_date(row[21])
                        pass_developer = row[22]
                        gidr_constr = row[23]
                        gidr_number = row[24]
                        gidr_coordinates = row[25]

                        LandPlot.objects.create(
                            cadastr_number=cadastr_number,
                            location=location,
                            area=area,
                            coordinates = coordinates,
                            status = get_status_key(status),
                            owner_name=owner_name,
                            category=get_category_key(category),
                            destination=get_destination_key(destination),
                            rent_start=rent_start,
                            rent_end=rent_end,
                            interest=interest,
                            assessment=assessment,
                            land=get_land_key(land),
                            registered=get_registered_key(registered),
                            notes=notes,
                            pass_number=pass_number,
                            pass_name=pass_name,
                            pass_area=pass_area,
                            pass_volume=pass_volume,
                            pass_depth=pass_depth,
                            date_approval=date_approval,
                            pass_developer=pass_developer,
                            gidr_constr = get_gidr_key(gidr_constr),
                            gidr_number = gidr_number,
                            gidr_coordinates = gidr_coordinates
                        )
                    except Exception as row_error:
                        errors.append(f"Рядок {i}: {str(row_error)}")

                if errors:
                    error_message = "Імпорт завершено з помилками:<br>" + "<br>".join(errors)
                    messages.warning(request, format_html(error_message))
                else:
                    messages.success(request, "Дані успішно імпортовано.")

                return redirect("import_landplots_excel")
            except Exception as e:
                messages.error(request, f"Помилка відкриття файлу: {str(e)}")
    else:
        form = ExcelImportForm()
    return render(request, "land/import_excel.html", {"form": form})


# Пагінація
def landplot_list(request):
    landplots = LandPlot.objects.all().order_by('id')  # Можна додати фільтрацію

    paginator = Paginator(landplots, 50)  # 20 записів на сторінку
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'land/landplot_list.html', {'page_obj': page_obj})

def your_view(request):
    leases = LandPlot.objects.all()
    paginator = Paginator(leases, 50)  # по 10 записів на сторінку
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    return render(request, "your_template.html", {"page_obj": page_obj})

# автопідказка при вводі користувача
def owner_autocomplete(request):
    if 'term' in request.GET:
        search_term = request.GET.get('term')
        print(f"Search term: {search_term}")
        # Фільтруємо унікальні значення вже на сервері
        qs = LandPlot.objects.filter(owner_name__icontains=search_term).values_list('owner_name', flat=True).distinct()
        return JsonResponse(list(qs), safe=False)
    return JsonResponse([], safe=False)

# автопідказка при вводі місцерозташування
def location_autocomplete(request):
    if 'term' in request.GET:
        search_term = request.GET.get('term')
        locations = LandPlot.objects.filter(location__icontains=search_term).values_list('location', flat=True).distinct()
        return JsonResponse(list(locations), safe=False)
    return JsonResponse([], safe=False)

# автопідказка при вводі кадастрового номера
def cadastr_autocomplete(request):
    if 'term' in request.GET:
        search_term = request.GET.get('term')
        cadastr = LandPlot.objects.filter(cadastr_number__icontains=search_term).values_list('cadastr_number', flat=True).distinct()
        return JsonResponse(list(cadastr), safe=False)
    return JsonResponse([], safe=False)

# погода
def weather_view(request):
    city = request.GET.get('city', 'Жашків')  # можна передавати місто через URL
    api_key = '8d2f86ddab8282673d32ced9d3f19ebd'  # встав свій API ключ тут
    url = f'https://api.openweathermap.org/data/2.5/weather?q={city}&appid={api_key}&units=metric&lang=ua'

    response = requests.get(url)
    data = response.json()

    weather = {}
    if response.status_code == 200:
        weather = {
            'city': city,
            'temperature': data['main']['temp'],
            'feels_like': data['main']['feels_like'],
            'humidity': data['main']['humidity'],
            'pressure': data['main']['pressure'],
            'wind_speed': data['wind']['speed'],
            'description': data['weather'][0]['description'],
            'icon': data['weather'][0]['icon'],
            'clouds': data['clouds']['all'],
        }
    else:
        weather['error'] = 'Місто не знайдено'

    return render(request, 'land/weather.html', {'weather': weather, 'title': 'Погода',})

def weather_context(request):
    city = 'Жашків'  # Можеш зробити динамічним, якщо хочеш
    api_key = '8d2f86ddab8282673d32ced9d3f19ebd'
    url = f'https://api.openweathermap.org/data/2.5/weather?q={city}&appid={api_key}&units=metric&lang=ua'

    weather = None
    try:
        response = requests.get(url)
        data = response.json()
        if response.status_code == 200:
            weather = {
            'city': city,
            'temperature': data['main']['temp'],
            'feels_like': data['main']['feels_like'],
            'humidity': data['main']['humidity'],
            'pressure': data['main']['pressure'],
            'wind_speed': data['wind']['speed'],
            'description': data['weather'][0]['description'],
            'icon': data['weather'][0]['icon'],
            'clouds': data['clouds']['all'],
            }
    except Exception:
        weather = None

    return {'weather': weather}


def all_land(request):
    land_plots = LandPlot.objects.all()
    search_query = request.GET.get('search', '')

    if search_query:
        land_plots = land_plots.filter(
            Q(cadastr_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(area__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(interest__icontains=search_query) |
            Q(assessment__icontains=search_query) |
            Q(rent_start__icontains=search_query) |
            Q (rent_end__icontains=search_query)
        )
    land_filter = LeaseFilterForm(request.GET.copy(), queryset=land_plots)
    filtered_qs = land_filter.qs.order_by('id')

    paginator = Paginator(filtered_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('land/zem_lease_results.html', {
            'leases': page_obj,
            'page_obj': page_obj,
        })
        return JsonResponse({'html': html})

    return render(request, 'land/land.html', {
        'title': 'Земельні ділнки',
        'filter': land_filter,
        'leases': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
    })

# експорт земельна ділянка
def land_export(request):
    print("=== Виклик land_export ===")
    print(f"GET-параметри в land_export: {request.GET}")
    print(f"POST-дані: {request.POST}")

    # Обробка параметрів фільтра
    get_params = request.GET.copy()
    if 'geom_filter' in get_params and get_params['geom_filter'] == '':
        get_params['geom_filter'] = None

    filter = LeaseFilterForm(get_params, queryset=LandPlot.objects.all())
    if not filter.is_valid():
        print(f"Помилки форми: {filter.errors}")
    landplots = filter.qs
    print(f"Кількість відфільтрованих ділянок: {landplots.count()}")
    print(f"Фільтровані ділянки (ID): {[plot.id for plot in landplots]}")
    print(f"Усі ділянки в базі: {LandPlot.objects.count()}")
    print(f"Ділянки з geom__isnull=True: {LandPlot.objects.filter(geom__isnull=True).count()}")
    print(f"Ділянки з geom__isnull=False: {LandPlot.objects.filter(geom__isnull=False).count()}")

    # Визначення всіх можливих колонок та їх відповідність полям моделі
    all_headers = [
        ("№", lambda plot, num: num),
        ("Кадастровий номер", lambda plot, num: plot.cadastr_number),
        ("Населений пункт", lambda plot, num: plot.location),
        ("Площа, га", lambda plot, num: plot.area),
        ("Користувач", lambda plot, num: plot.owner_name),
        ("Категорія", lambda plot, num: plot.get_category_display()),
        ("Цільове призначення", lambda plot, num: plot.get_destination_display()),
        ("Номер рішення", lambda plot, num: plot.decision_number),
        ("Дата рішення", lambda plot, num: plot.decision_date.strftime('%d.%m.%Y') if plot.decision_date else ''),
        ("Дата реєстрації", lambda plot, num: plot.rent_start.strftime('%d.%m.%Y') if plot.rent_start else ''),
        ("Термін дії", lambda plot, num: plot.rent_end.strftime('%d.%m.%Y') if plot.rent_end else ''),
        ("Відсоток", lambda plot, num: plot.interest),
        ("Нормативна оцінка", lambda plot, num: plot.assessment),
        ("Угіддя", lambda plot, num: plot.get_land_display()),
        ("Примітка", lambda plot, num: plot.notes),
        ("Суборендар", lambda plot, num: plot.sub_owner_name),
        ("Дата реєстрації (суборенда)", lambda plot, num: plot.sub_rent_start.strftime('%d.%m.%Y') if plot.sub_rent_start else ''),
        ("Термін дії (суборенда)", lambda plot, num: plot.sub_rent_end.strftime('%d.%m.%Y') if plot.sub_rent_end else ''),
        ("№ реєстру", lambda plot, num: plot.pass_number),
        ("Назва паспорту", lambda plot, num: plot.pass_name),
        ("Площа, га (паспорт)", lambda plot, num: plot.pass_area),
        ("Обєм при НПР", lambda plot, num: plot.pass_volume),
        ("Глибина, м", lambda plot, num: plot.pass_depth),
        ("Дата погодження", lambda plot, num: plot.date_approval.strftime('%d.%m.%Y') if plot.date_approval else ''),
        ("Розробник", lambda plot, num: plot.pass_developer),
        ("Гідроспоруда", lambda plot, num: plot.get_gidr_constr_display()),
        ("Номер акту", lambda plot, num: plot.gidr_number),
        ("Координати гідроспоруди", lambda plot, num: plot.gidr_coordinates),
    ]

    # Перевірка, які колонки мають дані
    headers = []
    header_functions = []
    for header, func in all_headers:
        # Перевіряємо, чи є ненульові/непорожні значення для цього поля
        has_data = False
        for plot in landplots:
            value = func(plot, 1)  # Використовуємо номер 1 для тесту
            if value not in (None, '', 0, 'None'):
                has_data = True
                break
        if has_data or header == "№":  # Завжди включаємо колонку "№"
            headers.append(header)
            header_functions.append(func)

    # Створення Excel-файлу
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LandPlots"

    # Додавання заголовків
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Заповнення рядків
    for num, plot in enumerate(landplots, start=1):
        row = [func(plot, num) for func in header_functions]
        sheet.append(row)

    # Налаштування ширини колонок
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Формування відповіді
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=landplots.xlsx'
    workbook.save(response)
    return response


def water(request):
    water_leases = LandPlot.objects.filter(category='H')
    search_query = request.GET.get('search', '')

    if search_query:
        water_leases = water_leases.filter(
            Q(cadastr_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(area__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(interest__icontains=search_query) |
            Q(assessment__icontains=search_query) |
            Q(rent_start__icontains=search_query) |
            Q (rent_end__icontains=search_query)
        )
    water_filter = LeaseFilterForm(request.GET, queryset=water_leases)
    filtered_qs = water_filter.qs

    paginator = Paginator(filtered_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('land/zem_lease_results.html', {
            'leases': page_obj,
            'page_obj': page_obj,
        })
        return JsonResponse({'html': html})

    return render(request, 'land/water.html', {
        'title': 'Землі водного фонду',
        'filter': water_filter,
        'leases': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
    })

def water_export_excel(request):
    # Створюємо новий Excel файл
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Water"

    # Заголовки
    headers = ["№", "Кадастровий номер", "Площа, га", "Населений пункт", "Цільове призначення", "Угіддя", "Координати", "Користувач",
               "Дата реєстрації", "Термін дії", "Відсоток", "Нормативна оцінка", "Примітка", "№ реєстру", "Назва паспорту",
               "Площа, га", "Обєм при НПР", "Глибина, м", "Дата погодження", "Розробник", "Гідроспоруда", "номер акту", "Координати гідроспоруди"]
    sheet.append(headers)
    # Жирний стиль для заголовків
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Дані з моделі
    waters = LandPlot.objects.filter(category='H')
    cadastr_number = request.GET.get("cadastr_number")
    if cadastr_number:
        waters = waters.filter(cadastr_number=cadastr_number)

    area = request.GET.get("area")
    if area:
        try:
            waters = waters.filter(area=float(area))
        except ValueError:
            pass

    location = request.GET.get("location")
    if location:
        waters = waters.filter(location__icontains=location)

    destination = request.GET.get("destination")
    if destination:
        waters = waters.filter(destination=destination)

    land = request.GET.get("land")
    if land:
        waters = waters.filter(land=land)

    coordinates = request.GET.get("coordinates")
    if coordinates:
        try:
            waters = waters.filter(coordinates=float(coordinates))
        except ValueError:
            pass

    owner_name = request.GET.get("owner_name")
    if owner_name:
        waters = waters.filter(owner_name__icontains=owner_name)

    rent_start = request.GET.get("rent_start")
    if rent_start:
        try:
            date_start = datetime.strptime(rent_start, '%Y-%m-%d').date()
            waters = waters.filter(rent_start=date_start)
        except ValueError:
            pass

    rent_end = request.GET.get("rent_end")
    if rent_end:
        try:
            date_end = datetime.strptime(rent_end , '%Y-%m-%d').date()
            waters = waters.filter(rent_end =date_end)
        except ValueError:
            pass

    interest = request.GET.get("interest")
    if interest:
        try:
            waters = waters.filter(interest=float(interest))
        except ValueError:
            pass

    assessment = request.GET.get("assessment")
    if assessment:
        try:
            waters = waters.filter(assessment=float(assessment))
        except ValueError:
            pass

    date_approval = request.GET.get("date_approval")
    if date_approval:
        try:
            date_approval = datetime.strptime(rent_end , '%Y-%m-%d').date()
            waters = waters.filter(date_approval =date_approval)
        except ValueError:
            pass

    gidr_constr = request.GET.get("gidr_constr")
    if gidr_constr:
        try:
            waters = waters.filter(gidr_constr__icontains=gidr_constr)
        except ValueError:
            pass

    gidr_number = request.GET.get("gidr_number")
    if gidr_number:
        try:
            waters = waters.filter(gidr_number__icontains=gidr_number)
        except ValueError:
            pass

    gidr_coordinates = request.GET.get("gidr_coordinates")
    if gidr_coordinates:
        try:
            waters = waters.filter(gidr_coordinates__icontains=gidr_coordinates)
        except ValueError:
            pass


    for num, plot in enumerate(waters, start=1):
        sheet.append([
            num,
            plot.cadastr_number,
            plot.area,
            plot.location,
            plot.get_destination_display(),
            plot.get_land_display(),
            plot.coordinates,
            plot.owner_name,
            plot.rent_start.strftime('%d.%m.%Y') if plot.rent_start else '',
            plot.rent_end.strftime('%d.%m.%Y') if plot.rent_end else '',
            plot.interest,
            plot.assessment,
            plot.notes,
            plot.pass_number,
            plot.pass_name,
            plot.pass_area,
            plot.pass_volume,
            plot.pass_depth,
            plot.date_approval.strftime('%d.%m.%Y') if plot.date_approval else '',
            plot.pass_developer,
            plot.get_gidr_constr_display(),
            plot.gidr_number,
            plot.gidr_coordinates
        ])
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
    # Відповідь
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=waters.xlsx'
    workbook.save(response)
    return response


# вільні земельні ділянки
def freeland(request):
    land_plots = LandPlot.objects.filter(status='1')
    search_query = request.GET.get('search', '')

    if search_query:
        land_plots = land_plots.filter(
            Q(cadastr_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(area__icontains=search_query)
        )
    land_filter = LeaseFilterForm(request.GET, queryset=land_plots)
    filtered_qs = land_filter.qs

    paginator = Paginator(filtered_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('land/zem_lease_results.html', {
            'leases': page_obj,
            'page_obj': page_obj,
        })
        return JsonResponse({'html': html})

    return render(request, 'land/freeland.html', {
        'title': 'Вільні земельні ділянки',
        'filter': land_filter,
        'leases': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
    })

# власність та постійне користування
def permanent_land(request):
    land_plots = LandPlot.objects.filter(status__in=['3', '4'])
    search_query = request.GET.get('search', '')

    if search_query:
        land_plots = land_plots.filter(
            Q(cadastr_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(area__icontains=search_query)  )
    land_filter = LeaseFilterForm(request.GET, queryset=land_plots)
    filtered_qs = land_filter.qs

    paginator = Paginator(filtered_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('land/zem_lease_results.html', {
            'leases': page_obj,
            'page_obj': page_obj,
        })
        return JsonResponse({'html': html})

    return render(request, 'land/permanent_land.html', {
        'title': 'Власність та постійне користування',
        'filter': land_filter,
        'leases': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
    })

# аукціони
def auction(request):
    land_plots = LandPlot.objects.filter(status='5')
    search_query = request.GET.get('search', '')

    if search_query:
        land_plots = land_plots.filter(
            Q(cadastr_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(area__icontains=search_query)
        )
    land_filter = LeaseFilterForm(request.GET, queryset=land_plots)
    filtered_qs = land_filter.qs

    paginator = Paginator(filtered_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('land/zem_lease_results.html', {
            'leases': page_obj,
            'page_obj': page_obj,
        })
        return JsonResponse({'html': html})

    return render(request, 'land/auction.html', {
        'title': 'Аукціони',
        'filter': land_filter,
        'leases': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
    })

# мапа Жашків
def map_view(request):
    return render(request, 'land/map.html', {'title': 'Мапа Жашкова'})

def landplots_geojson(request):
    bbox = request.GET.get('bbox')
    status = request.GET.get('status')
    category = request.GET.get('category')
    limit = int(request.GET.get('limit', 1000))
    offset = int(request.GET.get('offset', 0))
    try:
        queryset = LandPlot.objects.filter(geom__isnull=False)
        if bbox:
            try:
                minx, miny, maxx, maxy = map(float, bbox.split(','))
                bbox_poly = Polygon.from_bbox((minx, miny, maxx, maxy))
                queryset = queryset.filter(geom__intersects=bbox_poly)
            except ValueError as e:
                return JsonResponse({'error': f'Невірний формат bbox: {e}'}, status=400)

        if status and status != '':
            queryset = queryset.filter(status=status)
        if category and category != '':
            queryset = queryset.filter(category=category)

        print("Фільтр status:", status)
        print("Кількість ділянок:", queryset.count())

        queryset = queryset[offset:offset + limit]
        features = []
        for plot in queryset.annotate(geom_wgs=Transform('geom', 4326)):
            geom = json.loads(plot.geom_wgs.geojson) if plot.geom_wgs else None
            if not geom:
                continue
            features.append({
                "type": "Feature",
                "geometry": geom,
                "properties": {
                    "cadastr_number": plot.cadastr_number,
                    "area": float(plot.area) if plot.area else None,
                    "location": plot.location,
                    "destination": plot.get_destination_display() if plot.destination else None,
                    "owner_name": plot.owner_name,
                    "status": plot.status,
                    "status_display": plot.get_status_display() if plot.status else None,
                    "category": plot.category,
                    "category_display": plot.get_category_display() if plot.category else None,
                    "sub_owner_name": plot.sub_owner_name

                }
            })

        if not features:
            print("Ділянки не знайдено для status:", status)
            return JsonResponse({"type": "FeatureCollection", "features": []}, safe=False)

        return JsonResponse({"type": "FeatureCollection", "features": features}, safe=False)
    except Exception as e:
        print("Помилка сервера:", str(e))
        return JsonResponse({'error': f'Помилка сервера: {str(e)}'}, status=500)

@api_view(['GET'])
def search_landplot(request):
    cadastr = request.GET.get('cadastr')
    user = request.GET.get('user')
    limit = int(request.GET.get('limit', 1000))
    offset = int(request.GET.get('offset', 0))
    status = request.GET.get('status')
    try:
        queryset = LandPlot.objects.filter(geom__isnull=False)
        if cadastr:
            # Нормалізація кадастрового номера
            cadastr = cadastr.strip().replace(' ', '')  # Видаляємо пробіли
            if not cadastr:  # Перевірка на порожній номер
                return JsonResponse({'error': 'Кадастровий номер не вказано'}, status=400)

            try:
                # Пошук із нечутливістю до регістру
                landplot = queryset.get(cadastr_number__iexact=cadastr)
                # Перевірка геометрії
                geom = json.loads(landplot.geom.transform(4326, clone=True).geojson) if landplot.geom else None
                if not geom:
                    return JsonResponse({'error': 'Геометрія ділянки відсутня'}, status=404)

                return JsonResponse({
                    "type": "Feature",
                    "geometry": geom,
                    "properties": {
                        "cadastr_number": landplot.cadastr_number,
                        "area": float(landplot.area) if landplot.area else None,
                        "location": landplot.location,
                        "destination": landplot.get_destination_display() if landplot.destination else None,
                        "owner_name": landplot.owner_name,
                        "status": landplot.status,
                        "status_display": landplot.get_status_display() if landplot.status else None,
                        "category": landplot.category,
                        "category_display": landplot.get_category_display() if landplot.category else None
                    }
                }, safe=False)
            except LandPlot.DoesNotExist:
                return JsonResponse({'error': f'Ділянка з кадастровим номером {cadastr} не знайдена'}, status=404)
            except Exception as e:
                return JsonResponse({'error': f'Помилка пошуку: {str(e)}'}, status=500)
        elif user:
            queryset = queryset.filter(owner_name__icontains=user)[offset:offset + limit]
            features = []
            for land in queryset.annotate(geom_wgs=Transform('geom', 4326)):
                geom = json.loads(land.geom_wgs.geojson) if land.geom else None
                if not geom:
                    continue
                features.append({
                    "type": "Feature",
                    "geometry": geom,
                    "properties": {
                        "cadastr_number": land.cadastr_number,
                        "area": float(land.area) if land.area else None,
                        "location": land.location,
                        "destination": land.get_destination_display() if land.destination else None,
                        "owner_name": land.owner_name,
                        "sub_owner_name": land.sub_owner_name
                    }
                })
            if not features:
                return JsonResponse({'error': 'Ділянки за цим користувачем не знайдено'}, status=404)
            return JsonResponse({"type": "FeatureCollection", "features": features}, safe=False)
        elif status:
                queryset = queryset.filter(status=status)
                print("Фільтр status:", status)
                print("Кількість ділянок:", queryset.count())

                serializer = Serializer()
                geojson_data = serializer.serialize(queryset, fields=(
                    'cadastr_number', 'area', 'location', 'category', 'category_display',
                    'destination', 'owner_name', 'status', 'status_display'
                ))
                return JsonResponse(json.loads(geojson_data), safe=False)
        return JsonResponse({'error': 'Невірні параметри'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Помилка сервера: {str(e)}'}, status=500)

def status_layer_view(request):
    bbox = request.GET.get('bbox')
    limit = int(request.GET.get('limit', 1000))
    offset = int(request.GET.get('offset', 0))
    status = request.GET.get('status')
    try:
        queryset = LandPlot.objects.filter(geom__isnull=False)
        if bbox:
            try:
                minx, miny, maxx, maxy = map(float, bbox.split(','))
                bbox_poly = Polygon.from_bbox((minx, miny, maxx, maxy))
                queryset = queryset.filter(geom__intersects=bbox_poly)
            except ValueError as e:
                return JsonResponse({'error': f'Невірний формат bbox: {str(e)}'}, status=400)
        if status and status != '':
            queryset = queryset.filter(status=status)
        queryset = queryset[offset:offset + limit]
        features = []
        for land in queryset.annotate(geom_wgs=Transform('geom', 4326)):
            geom = json.loads(land.geom_wgs.geojson) if land.geom_wgs else None
            if not geom:
                continue
            features.append({
                "type": "Feature",
                "geometry": geom,
                "properties": {
                    "cadastr_number": land.cadastr_number or None,
                    "status": land.status or None,
                    "status_display": land.get_status_display() if land.status else None,
                    "area": float(land.area) if land.area else None,
                    "location": land.location or None,
                    "destination": land.get_destination_display() if land.destination else None,
                    "owner_name": land.owner_name or None,
                    "category": land.category or None,
                    "category_display": land.get_category_display() if land.category else None
                }
            })
        if not features:
            return JsonResponse({"type": "FeatureCollection", "features": []}, safe=False)
        return JsonResponse({"type": "FeatureCollection", "features": features}, safe=False)
    except Exception as e:
        return JsonResponse({'error': f'Помилка сервера: {str(e)}'}, status=500)




def category_layer_view(request):
    bbox = request.GET.get('bbox')
    limit = int(request.GET.get('limit', 1000))
    offset = int(request.GET.get('offset', 0))

    try:
        queryset = LandPlot.objects.filter(geom__isnull=False)
        if bbox:
            try:
                minx, miny, maxx, maxy = map(float, bbox.split(','))
                bbox_poly = Polygon.from_bbox((minx, miny, maxx, maxy))
                queryset = queryset.filter(geom__intersects=bbox_poly)
            except ValueError as e:
                return JsonResponse({'error': f'Невірний формат bbox: {e}'}, status=400)

        queryset = queryset[offset:offset + limit]
        features = []
        for land in queryset.annotate(geom_wgs=Transform('geom', 4326)):
            geom = json.loads(land.geom_wgs.geojson) if land.geom else None
            if not geom:
                continue
            features.append({
                "type": "Feature",
                "geometry": geom,
                "properties": {
                    "cadastr_number": land.cadastr_number,
                    "category": land.category,
                    "category_display": land.get_category_display(),
                    "area": float(land.area) if land.area else None,
                    "location": land.location,
                    "destination": land.get_destination_display() if land.destination else None,
                    "owner_name": land.owner_name
                }
            })
        return JsonResponse({"type": "FeatureCollection", "features": features}, safe=False)
    except Exception as e:
        return JsonResponse({'error': f'Помилка сервера: {str(e)}'}, status=500)

